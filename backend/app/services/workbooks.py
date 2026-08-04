"""Safe, domain-neutral spreadsheet extraction and row chunk generation."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, time
from decimal import Decimal
from json import dumps
from pathlib import Path
import csv
import math
import re

from app.services.document_loader import DocumentParseError


STRUCTURED_INDEX_VERSION = "structured-workbook-v2"


@dataclass
class WorkbookRow:
    row_number: int
    values: dict[str, object]


@dataclass
class WorkbookSheet:
    name: str
    state: str
    status: str
    header_row: int | None = None
    headers: list[str] = field(default_factory=list)
    rows: list[WorkbookRow] = field(default_factory=list)
    error: str | None = None


@dataclass
class WorkbookData:
    sheets: list[WorkbookSheet]

    @property
    def non_empty_sheets(self) -> list[WorkbookSheet]:
        return [sheet for sheet in self.sheets if sheet.status == "processed"]

    @property
    def skipped_sheets(self) -> list[str]:
        return [sheet.name for sheet in self.sheets if sheet.status == "empty"]

    @property
    def failed_sheets(self) -> list[str]:
        return [sheet.name for sheet in self.sheets if sheet.status == "failed"]


def _normalized_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            return None
        return int(value) if value.is_integer() else value
    if isinstance(value, (str, int, bool)):
        return value
    return str(value)


def _xlsx_value(cell) -> object:
    value = _normalized_value(cell.value)
    if isinstance(value, int) and value >= 0:
        number_format = str(getattr(cell, "number_format", "") or "")
        if re.fullmatch(r"0+", number_format) and len(number_format) > len(str(value)):
            return str(value).zfill(len(number_format))
    return value


def _is_nonempty(value: object) -> bool:
    return value is not None and (not isinstance(value, str) or bool(value.strip()))


def _detect_header(rows: list[tuple[int, list[object]]]) -> int:
    """Choose a probable header from the leading rows without domain assumptions."""
    candidates = rows[:20]
    best_index = 0
    best_score = float("-inf")
    for index, (_, values) in enumerate(candidates):
        present = [value for value in values if _is_nonempty(value)]
        if not present:
            continue
        text_count = sum(isinstance(value, str) for value in present)
        distinct = len({str(value).strip().casefold() for value in present})
        following_width = 0
        if index + 1 < len(rows):
            following_width = sum(_is_nonempty(value) for value in rows[index + 1][1])
        score = len(present) * 3 + text_count * 2 + distinct + min(following_width, len(present))
        if len(present) == 1 and len(rows) > index + 1:
            score -= 4
        score -= index
        if score > best_score:
            best_index, best_score = index, score
    return best_index


def _headers(header_rows: list[list[object]], width: int) -> list[str]:
    """Build unique headers, preserving useful labels from stacked header rows."""
    output: list[str] = []
    used: dict[str, int] = {}
    for index in range(width):
        parts: list[str] = []
        for values in header_rows:
            raw = values[index] if index < len(values) else None
            if not _is_nonempty(raw):
                continue
            part = str(raw).strip()
            if part.casefold() not in {value.casefold() for value in parts}:
                parts.append(part)
        if len(parts) > 1:
            descriptive = [part for part in parts if len(part) > 1]
            if descriptive:
                parts = descriptive
        base = " / ".join(parts) if parts else f"Column {index + 1}"
        count = used.get(base.casefold(), 0) + 1
        used[base.casefold()] = count
        output.append(base if count == 1 else f"{base} ({count})")
    return output


def _header_layers(
    rows: list[tuple[int, list[object]]],
    header_index: int,
) -> list[tuple[int, list[object]]]:
    """Include one adjacent, substantial row when a sheet uses stacked headers."""
    layers = [rows[header_index]]
    if header_index == 0:
        return layers
    previous = rows[header_index - 1]
    previous_present = sum(_is_nonempty(value) for value in previous[1])
    current_present = sum(_is_nonempty(value) for value in rows[header_index][1])
    if (
        previous[0] + 1 == rows[header_index][0]
        and previous_present >= 2
        and current_present >= 2
        and previous_present >= current_present * 0.5
    ):
        layers.insert(0, previous)
    return layers


def _normalize_attendance_pairs(
    headers: list[str],
    rows: list[WorkbookRow],
) -> tuple[list[str], list[WorkbookRow]]:
    """Split paired employee name/number cells across adjacent IN and OUT rows."""
    entity_header = next(
        (
            header
            for header in headers
            if "employee" in header.casefold() and "name" in header.casefold()
        ),
        None,
    )
    if entity_header is None:
        return headers, rows
    marker_header = next(
        (
            header
            for header in headers
            if {"in", "out"} <= {
                str(row.values.get(header) or "").strip().casefold()
                for row in rows
            }
        ),
        None,
    )
    if marker_header is None:
        return headers, rows

    normalized_rows = [
        WorkbookRow(
            row_number=row.row_number,
            values={
                "Employee Name": None,
                "Employee Number": None,
                **{
                    header: value
                    for header, value in row.values.items()
                    if header != entity_header
                },
            },
        )
        for row in rows
    ]
    group_start: int | None = None
    employee_name: object = None
    employee_number: object = None
    for index, row in enumerate(rows):
        marker = str(row.values.get(marker_header) or "").strip().casefold()
        if marker == "in":
            group_start = index
            employee_name = row.values.get(entity_header)
            employee_number = None
        elif marker == "out" and group_start is not None:
            employee_number = row.values.get(entity_header)
            for grouped_index in range(group_start, index):
                normalized_rows[grouped_index].values["Employee Number"] = employee_number
        if group_start is not None:
            normalized_rows[index].values["Employee Name"] = employee_name
            normalized_rows[index].values["Employee Number"] = employee_number

    normalized_headers = [
        "Employee Name",
        "Employee Number",
        *(header for header in headers if header != entity_header),
    ]
    return normalized_headers, normalized_rows


def _make_sheet(name: str, state: str, rows: list[tuple[int, list[object]]]) -> WorkbookSheet:
    nonempty = [
        (number, values)
        for number, values in rows
        if any(_is_nonempty(value) for value in values)
    ]
    if not nonempty:
        return WorkbookSheet(name=name, state=state, status="empty")

    header_index = _detect_header(nonempty)
    header_layers = _header_layers(nonempty, header_index)
    header_number = header_layers[0][0]
    data_rows = nonempty[header_index + 1 :]
    width = max([
        *(len(values) for _, values in header_layers),
        *(len(values) for _, values in data_rows),
    ])
    headers = _headers([values for _, values in header_layers], width)
    structured = []
    for row_number, values in data_rows:
        row = {
            header: values[index] if index < len(values) else None
            for index, header in enumerate(headers)
        }
        if any(_is_nonempty(value) for value in row.values()):
            structured.append(WorkbookRow(row_number=row_number, values=row))
    headers, structured = _normalize_attendance_pairs(headers, structured)
    return WorkbookSheet(
        name=name,
        state=state,
        status="processed",
        header_row=header_number,
        headers=headers,
        rows=structured,
    )


def _extract_xlsx(
    path: Path, include_hidden: bool, include_very_hidden: bool
) -> WorkbookData:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    sheets: list[WorkbookSheet] = []
    try:
        for worksheet in workbook.worksheets:
            state = str(worksheet.sheet_state)
            if (
                state == "hidden" and not include_hidden
                or state == "veryHidden" and not include_very_hidden
            ):
                sheets.append(WorkbookSheet(worksheet.title, state, "disabled"))
                continue
            try:
                rows = [
                    (index, [_xlsx_value(cell) for cell in row])
                    for index, row in enumerate(worksheet.iter_rows(), start=1)
                ]
                sheets.append(_make_sheet(worksheet.title, state, rows))
            except Exception:
                sheets.append(
                    WorkbookSheet(
                        worksheet.title,
                        state,
                        "failed",
                        error="This worksheet could not be read.",
                    )
                )
    finally:
        workbook.close()
    return WorkbookData(sheets)


def _extract_xls(
    path: Path, include_hidden: bool, include_very_hidden: bool
) -> WorkbookData:
    import xlrd

    workbook = xlrd.open_workbook(str(path), on_demand=True)
    sheets: list[WorkbookSheet] = []
    try:
        for index in range(workbook.nsheets):
            worksheet = workbook.sheet_by_index(index)
            visibility = int(workbook.sheet_visibility[index])
            state = (
                "visible" if visibility == 0
                else "hidden" if visibility == 1
                else "veryHidden"
            )
            if (
                visibility == 1 and not include_hidden
                or visibility >= 2 and not include_very_hidden
            ):
                sheets.append(WorkbookSheet(worksheet.name, state, "disabled"))
                continue
            try:
                rows: list[tuple[int, list[object]]] = []
                for row_index in range(worksheet.nrows):
                    values: list[object] = []
                    for column_index in range(worksheet.ncols):
                        cell = worksheet.cell(row_index, column_index)
                        value: object = cell.value
                        if cell.ctype == xlrd.XL_CELL_DATE:
                            value = xlrd.xldate_as_datetime(value, workbook.datemode)
                        elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                            value = bool(value)
                        elif cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
                            value = None
                        values.append(_normalized_value(value))
                    rows.append((row_index + 1, values))
                sheets.append(_make_sheet(worksheet.name, state, rows))
            except Exception:
                sheets.append(
                    WorkbookSheet(
                        worksheet.name,
                        state,
                        "failed",
                        error="This worksheet could not be read.",
                    )
                )
    finally:
        workbook.release_resources()
    return WorkbookData(sheets)


def _extract_csv(path: Path) -> WorkbookData:
    with path.open(
        "r",
        encoding="utf-8-sig",
        errors="replace",
        newline="",
    ) as handle:
        sample = handle.read(8192)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        rows = [
            (
                row_number,
                [_normalized_value(value.strip()) for value in row],
            )
            for row_number, row in enumerate(
                csv.reader(handle, dialect),
                start=1,
            )
        ]
    return WorkbookData([
        _make_sheet("CSV", "visible", rows)
    ])


def extract_workbook(
    path: Path,
    include_hidden: bool = True,
    include_very_hidden: bool = False,
) -> WorkbookData:
    """Read all configured sheets without evaluating formulas, links, or macros."""
    try:
        if path.suffix.lower() == ".xlsx":
            workbook = _extract_xlsx(path, include_hidden, include_very_hidden)
        elif path.suffix.lower() == ".xls":
            workbook = _extract_xls(path, include_hidden, include_very_hidden)
        elif path.suffix.lower() == ".csv":
            workbook = _extract_csv(path)
        else:
            raise DocumentParseError("Unsupported spreadsheet type.")
    except DocumentParseError:
        raise
    except Exception as error:
        label = (
            "CSV table"
            if path.suffix.lower() == ".csv"
            else "legacy Excel workbook"
            if path.suffix.lower() == ".xls"
            else "Excel workbook"
        )
        raise DocumentParseError(f"The {label} could not be read.") from error
    if not workbook.non_empty_sheets:
        failed = ", ".join(workbook.failed_sheets)
        detail = f" Affected sheets: {failed}." if failed else ""
        raise DocumentParseError(f"The workbook contains no readable non-empty worksheets.{detail}")
    return workbook


def _plain_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = " ".join(text.replace("\x00", "").split())
    if text.startswith(("=", "+", "-", "@")):
        text = f"'{text}"
    return text.replace("|", r"\|")


def _infer_type(values: list[object]) -> str:
    """Infer a general column type from observed cell values."""
    present = [value for value in values if _is_nonempty(value)]
    if not present:
        return "text"
    numbers = sum(isinstance(value, (int, float, Decimal)) and not isinstance(value, bool) for value in present)
    date_like = sum(isinstance(value, (date, datetime, time)) for value in present)
    strings = [str(value).strip() for value in present if isinstance(value, str)]
    if numbers / len(present) >= 0.8:
        return "number"
    if date_like / len(present) >= 0.8:
        return "date"
    unique_ratio = len({str(value).strip().casefold() for value in present}) / max(len(present), 1)
    if strings and all(re.fullmatch(r"[A-Za-z0-9_.:/#-]+", value) for value in strings) and unique_ratio > 0.8:
        return "identifier"
    if unique_ratio <= 0.5:
        return "category"
    return "text"


def workbook_schema(sheet: WorkbookSheet) -> dict[str, object]:
    """Build a compact schema for planning structured answers later."""
    columns = []
    for header in sheet.headers:
        values = [row.values.get(header) for row in sheet.rows]
        present = [value for value in values if _is_nonempty(value)]
        columns.append({
            "name": header,
            "type": _infer_type(values),
            "non_empty_count": len(present),
            "unique_count": len({str(value).strip().casefold() for value in present}),
        })
    return {
        "sheet": sheet.name,
        "header_row": sheet.header_row,
        "row_count": len(sheet.rows),
        "columns": columns,
    }


def workbook_chunks(workbook: WorkbookData, filename: str) -> list[tuple[str, str, int | None]]:
    """Create row-oriented chunks, each carrying workbook and sheet provenance."""
    chunks: list[tuple[str, str, int | None]] = []
    for sheet in workbook.non_empty_sheets:
        if not sheet.rows:
            header_text = " | ".join(_plain_text(header) for header in sheet.headers)
            chunks.append((
                f"Workbook: {_plain_text(filename)} | Sheet: {_plain_text(sheet.name)} | "
                f"Headers: {header_text}",
                sheet.name,
                sheet.header_row,
            ))
            continue
        for row in sheet.rows:
            fields = " | ".join(
                f"{_plain_text(header)}: {_plain_text(value)}"
                for header, value in row.values.items()
                if _is_nonempty(value)
            )
            chunks.append((
                f"Workbook: {_plain_text(filename)} | Sheet: {_plain_text(sheet.name)} | "
                f"Row: {row.row_number} | {fields}",
                sheet.name,
                row.row_number,
            ))
    return chunks


def workbook_text(workbook: WorkbookData, filename: str) -> str:
    metadata = {
        "empty_sheets": workbook.skipped_sheets,
        "failed_sheets": workbook.failed_sheets,
    }
    chunks = [text for text, _, _ in workbook_chunks(workbook, filename)]
    return "\n".join([f"Workbook processing metadata: {dumps(metadata)}", *chunks])
