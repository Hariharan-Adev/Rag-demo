"""Domain-neutral, schema-driven workbook RAG tests."""

from __future__ import annotations

import asyncio
from contextlib import ExitStack
from datetime import time
from io import BytesIO
from pathlib import Path
import tempfile
import unittest
from unittest.mock import patch

from fastapi import UploadFile
from openpyxl import Workbook
from starlette.requests import Request

from app import database
from app.routes import upload
from app.services import structured_ingestion, vector_search, vector_store
from app.services.chat_context import save_grounded_context
from app.services.rag_service import answer_question
from app.services.vector_store import reset_vector_store_for_tests
from app.services.workbook_analysis import (
    RowRecord,
    WorkbookScope,
    analyze_workbook_question,
    _answer_time_matrix,
    _column_score,
    _row_filters,
)
from app.services.workbooks import _make_sheet


ORG = "00000000-0000-4000-8000-000000000001"


def _request() -> Request:
    return Request({
        "type": "http",
        "method": "POST",
        "path": "/documents/upload",
        "headers": [],
        "client": ("test", 1),
    })


def _workbook_bytes(sheets: list[tuple[str, list[list[object]], str]]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, rows, state in sheets:
        sheet = workbook.create_sheet(name)
        sheet.sheet_state = state
        for row in rows:
            sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


class WorkbookDomainNeutralTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        root = Path(self.temporary.name)
        self.database_path = root / "test.db"
        self.upload_path = root / "uploads"
        self.stack = ExitStack()
        self.stack.enter_context(patch.object(database, "DATABASE_PATH", self.database_path))
        self.stack.enter_context(patch.object(database, "UPLOAD_DIRECTORY", self.upload_path))
        self.stack.enter_context(patch.object(vector_store.settings, "vector_store", "sqlite"))
        self.stack.enter_context(patch.object(vector_store.settings, "vector_store_provider", "sqlite"))
        self.stack.enter_context(patch.object(vector_store.settings, "qdrant_local_path", ""))
        self.stack.enter_context(patch.object(upload, "UPLOAD_DIRECTORY", self.upload_path))
        self.stack.enter_context(patch.object(upload, "enforce_request_limit", lambda *args, **kwargs: None))
        self.stack.enter_context(patch.object(upload, "log_audit_event", lambda **kwargs: None))
        self.stack.enter_context(
            patch.object(
                upload,
                "create_embeddings",
                lambda chunks: [
                    ([1.0, 0.0] if "FINAL_ONLY" in chunk else [0.0, 1.0]) + [0.0] * 382
                    for chunk in chunks
                ],
            )
        )
        database.initialize_database()
        reset_vector_store_for_tests()
        with database.get_connection() as connection:
            connection.executemany(
                "INSERT INTO users (id, email, password_hash) VALUES (?, ?, 'hash')",
                [(1, "owner@example.com"), (2, "other@example.com")],
            )

    def tearDown(self) -> None:
        self.stack.close()
        self.temporary.cleanup()

    def upload_workbook(
        self,
        sheets: list[tuple[str, list[list[object]], str]],
        filename: str,
        owner_id: int = 1,
    ) -> dict[str, object]:
        file = UploadFile(file=BytesIO(_workbook_bytes(sheets)), filename=filename)
        return asyncio.run(upload._process_document_upload(_request(), file, {"id": owner_id}))

    def upload_text(self, text: str, filename: str, owner_id: int = 1) -> dict[str, object]:
        file = UploadFile(file=BytesIO(text.encode("utf-8")), filename=filename)
        return asyncio.run(upload._process_document_upload(_request(), file, {"id": owner_id}))

    def test_multi_sheet_tables_are_schema_indexed_and_counted(self) -> None:
        result = self.upload_workbook(
            [
                ("North", [["Item", "Units"], ["A-1", 4], ["A-2", 6]], "visible"),
                ("South", [["Item", "Units"], ["B-1", 5]], "visible"),
            ],
            "inventory.xlsx",
        )

        answer = analyze_workbook_question("How many records are there?", 1, document_id=int(result["document_id"]))

        self.assertEqual(len(result["workbook"]["processed_sheets"]), 2)
        self.assertIn("Count: 3", answer["answer"])
        self.assertEqual({source["source_location"]["sheet_name"] for source in answer["sources"]}, {"North", "South"})

    def test_filtered_total_uses_all_rows_not_vector_excerpt(self) -> None:
        result = self.upload_workbook(
            [
                ("Q1", [["Period", "Region", "Amount"], ["2026-01-01", "East", 10], ["2026-02-01", "East", 20]], "visible"),
                ("Q2", [["Period", "Region", "Amount"], ["2026-07-01", "East", 30], ["2026-07-15", "West", 40]], "visible"),
            ],
            "ledger.xlsx",
        )

        answer = analyze_workbook_question("What is the total amount for July?", 1, document_id=int(result["document_id"]))

        self.assertIn("70", answer["answer"])
        self.assertNotIn("100", answer["answer"])
        self.assertEqual(answer["sources"][0]["source_location"]["sheet_name"], "Q2")

    def test_automatic_document_selection_uses_schema_and_values(self) -> None:
        first = self.upload_workbook(
            [("Catalog", [["Code", "Label"], ["X1", "Alpha"]], "visible")],
            "catalog.xlsx",
        )
        second = self.upload_workbook(
            [("Ledger", [["Period", "Amount"], ["March", 25]], "visible")],
            "ledger.xlsx",
        )

        answer = analyze_workbook_question("What is the total amount for March?", 1)

        self.assertIn("25", answer["answer"])
        self.assertEqual(answer["sources"][0]["document_id"], int(second["document_id"]))
        self.assertNotEqual(answer["sources"][0]["document_id"], int(first["document_id"]))

    def test_common_short_words_are_not_cell_value_filters(self) -> None:
        rows = [
            RowRecord("Sheet1", 2, {"State": "IN", "Amount": 5}),
            RowRecord("Sheet1", 3, {"State": "OUT", "Amount": 7}),
        ]

        self.assertEqual(_row_filters(rows, "What is the total amount in July?"), {})
        self.assertEqual(_row_filters(rows, "Show state IN"), {"State": {"in"}})
        self.assertEqual(_row_filters(rows, "Show the in time"), {"State": {"in"}})

    def test_stacked_attendance_headers_and_in_rows_are_preserved(self) -> None:
        sheet = _make_sheet(
            "Attendance",
            "visible",
            [
                (1, ["Attendance report", *([None] * 11)]),
                (2, ["Employee Name & No", None, "Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Mon", "Tue"]),
                (3, ["A", None, "1st", "2nd", "3rd", "4th", "5th", "6th", "7th", "8th", "9th", "10th"]),
                (4, ["Alice", "IN", *([time(9, 0)] * 10)]),
                (5, [101, "OUT", *([time(18, 0)] * 10)]),
                (6, [None, "Double OT", *([time(1, 0)] * 10)]),
                (7, ["Bob", "IN", *([time(8, 55)] * 10)]),
                (8, [102, "OUT", *([time(17, 45)] * 10)]),
            ],
        )

        self.assertEqual(sheet.header_row, 2)
        self.assertEqual(sheet.headers[:2], ["Employee Name", "Employee Number"])
        self.assertEqual(sheet.headers[3], "Sun / 1st")
        self.assertEqual(sheet.rows[1].values["Employee Name"], "Alice")
        self.assertEqual(sheet.rows[1].values["Employee Number"], 101)
        self.assertEqual(sheet.rows[2].values["Employee Name"], "Alice")
        self.assertEqual(sheet.rows[2].values["Employee Number"], 101)
        self.assertEqual(
            _row_filters(
                [RowRecord(sheet.name, row.row_number, row.values) for row in sheet.rows],
                "Give me the in time of all employees",
            ),
            {"Column 2": {"in"}},
        )

    def test_single_character_and_substring_headers_do_not_create_relevance(self) -> None:
        self.assertEqual(_column_score("A", "How many products rejected?"), 0)
        self.assertEqual(_column_score("art", "What is the partial count?"), 0)
        self.assertGreater(_column_score("Product Count", "How many product records?"), 0)

    def test_unrelated_workbook_cannot_win_generic_count(self) -> None:
        self.upload_workbook(
            [("Log", [["A", "Column 2"], ["x", "present"], ["y", "absent"]], "visible")],
            "neutral.xlsx",
        )

        answer = analyze_workbook_question("How many products rejected?", 1)

        self.assertFalse(answer["grounded"])
        self.assertEqual(answer["sources"], [])

    def test_relevant_pdf_style_retrieval_wins_over_unrelated_workbook(self) -> None:
        self.upload_workbook(
            [("Log", [["A", "Column 2"], ["x", "present"], ["y", "absent"]], "visible")],
            "neutral.xlsx",
        )
        text = self.upload_text(
            "The uploaded file discusses product rejection evidence.",
            "quality-summary.txt",
        )
        with database.get_connection() as connection:
            version_id = connection.execute(
                "SELECT current_version_id FROM documents WHERE id = ?",
                (int(text["document_id"]),),
            ).fetchone()["current_version_id"]
        pdf_source = {
            "document_id": int(text["document_id"]),
            "version_id": int(version_id),
            "filename": "quality-summary.txt",
            "content": "The uploaded file discusses product rejection evidence.",
            "source_type": "text",
            "source_location": {"section": "body"},
            "score": 0.82,
        }

        with patch("app.services.rag_service.search_chunks", return_value=[pdf_source]), patch(
            "app.services.rag_service.generate_answer",
            return_value={
                "answer": "The file contains product rejection evidence (quality-summary.pdf, page 2).",
                "prompt_tokens": 100,
                "completion_tokens": 20,
            },
        ), patch("app.services.rag_service.reserve_groq_call"), patch(
            "app.services.rag_service.record_groq_tokens"
        ), patch("app.services.rag_service.log_audit_event"):
            answer = answer_question("What are the product rejections?", 1)

        self.assertEqual(answer["question_type"], "retrieval")
        self.assertEqual(answer["sources"][0]["filename"], "quality-summary.txt")

    def test_quantity_question_uses_sum_when_numeric_column_matches(self) -> None:
        result = self.upload_workbook(
            [("Stock", [["Item", "Units"], ["A", 2], ["B", 3]], "visible")],
            "stock.xlsx",
        )

        answer = analyze_workbook_question(
            "How many units are there?",
            1,
            document_id=int(result["document_id"]),
        )

        self.assertIn("Total Units: 5", answer["answer"])

    def test_status_filtering_uses_filtered_count(self) -> None:
        result = self.upload_workbook(
            [("Cases", [["Item", "Status"], ["A", "open"], ["B", "closed"]], "visible")],
            "cases.xlsx",
        )

        answer = analyze_workbook_question(
            "How many open items?",
            1,
            document_id=int(result["document_id"]),
        )

        self.assertIn("Count: 1", answer["answer"])
        self.assertEqual(answer["_context"]["filters"], {"Status": ["open"]})

    def test_grouped_result_is_deterministic_structured_evidence(self) -> None:
        result = self.upload_workbook(
            [("Data", [["Category", "Amount"], ["A", 3], ["B", 5], ["A", 7]], "visible")],
            "grouped.xlsx",
        )

        answer = analyze_workbook_question("Group amount by category", 1, document_id=int(result["document_id"]))

        self.assertIn("A: 10", answer["answer"])
        self.assertIn("B: 5", answer["answer"])
        self.assertTrue(answer["grounded"])

    def test_complete_structured_answer_wins_over_partial_vector_context(self) -> None:
        result = self.upload_workbook(
            [("Data", [["Period", "Amount"], ["April", 5], ["April", 15]], "visible")],
            "complete.xlsx",
        )
        partial = {
            "document_id": int(result["document_id"]),
            "version_id": 0,
            "filename": "complete.xlsx",
            "content": "Period: April\nAmount: 5",
            "source_type": "excel",
            "source_location": {"sheet_name": "Data", "row_start": 2, "row_end": 2},
            "score": 0.99,
        }

        with patch("app.services.rag_service.search_chunks", return_value=[partial]):
            answer = answer_question("What is the total amount for April?", 1)

        self.assertIn("20", answer["answer"])
        self.assertEqual(answer["question_type"], "structured_analysis")

    def test_low_score_retrieval_requires_overlap_and_keeps_sources_empty_when_unavailable(self) -> None:
        unrelated = {
            "document_id": 50,
            "version_id": 51,
            "filename": "notes.txt",
            "content": "Completely different content.",
            "source_type": "text",
            "source_location": {},
            "score": 0.1,
        }

        with patch("app.services.rag_service.has_structured_workbook", return_value=False), patch(
            "app.services.rag_service.is_analytical_question", return_value=False
        ), patch("app.services.rag_service.search_chunks", side_effect=[[], [unrelated]]), patch(
            "app.services.rag_service.generate_answer", side_effect=AssertionError("unrelated source must not be used")
        ), patch("app.services.rag_service.log_audit_event"):
            answer = answer_question("Find alpha beta", 1)

        self.assertEqual(answer["answer"], "Information not available in the uploaded files.")
        self.assertFalse(answer["grounded"])
        self.assertEqual(answer["sources"], [])

    def test_follow_up_uses_latest_grounded_context_within_same_chat(self) -> None:
        result = self.upload_workbook(
            [("Rows", [["Code", "Amount"], ["R1", 9], ["R2", 11]], "visible")],
            "rows.xlsx",
        )
        total = answer_question("What is the total amount?", 1, conversation_id="chat-a")
        follow_up = answer_question("list them", 1, conversation_id="chat-a")
        other_chat = answer_question("list them", 1, conversation_id="chat-b")
        other_user = answer_question("list them", 2, conversation_id="chat-a")

        self.assertIn("20", total["answer"])
        self.assertIn("9", follow_up["answer"])
        self.assertIn("11", follow_up["answer"])
        self.assertEqual({source["document_id"] for source in follow_up["sources"]}, {int(result["document_id"])})
        self.assertEqual(other_chat["sources"], [])
        self.assertEqual(other_user["sources"], [])

    def test_clear_topic_change_does_not_reuse_stale_context(self) -> None:
        result = self.upload_workbook(
            [("Rows", [["Code", "Amount"], ["R1", 9], ["R2", 11]], "visible")],
            "rows.xlsx",
        )
        answer_question("What is the total amount?", 1, conversation_id="chat-a")

        with patch("app.services.rag_service.search_chunks", return_value=[]), patch(
            "app.services.rag_service.log_audit_event"
        ):
            changed = answer_question("show invoices", 1, conversation_id="chat-a")

        self.assertEqual(int(result["document_id"]), 1)
        self.assertFalse(changed["grounded"])
        self.assertEqual(changed["sources"], [])

    def test_aggregate_citation_carries_complete_row_range(self) -> None:
        result = self.upload_workbook(
            [("Cases", [["Item", "Status"], ["A", "open"], ["B", "open"]], "visible")],
            "cases.xlsx",
        )

        answer = analyze_workbook_question(
            "How many open items?",
            1,
            document_id=int(result["document_id"]),
        )

        location = answer["sources"][0]["source_location"]
        self.assertEqual(location["row_start"], 2)
        self.assertEqual(location["row_end"], 3)

    def test_unavailable_and_acl_isolation_have_no_sources(self) -> None:
        result = self.upload_workbook(
            [("Private", [["Reference", "Value"], ["SECRET", 500]], "visible")],
            "private.xlsx",
        )

        unavailable = analyze_workbook_question("What is the total missing field?", 1, document_id=int(result["document_id"]))
        inaccessible = analyze_workbook_question("What is the total value?", 2, document_id=int(result["document_id"]))

        self.assertEqual(unavailable["sources"], [])
        self.assertFalse(unavailable["grounded"])
        self.assertEqual(inaccessible["sources"], [])
        self.assertNotIn("500", inaccessible["answer"])

    def test_source_plan_mismatch_is_not_saved_as_follow_up_context(self) -> None:
        save_grounded_context(
            owner_id=1,
            conversation_id="mismatch",
            question="older",
            result={
                "answer": "Grounded",
                "grounded": True,
                "sources": [{"document_id": 1, "version_id": 2, "filename": "a.xlsx"}],
                "_context": {"document_ids": [99], "version_ids": [2], "result_type": "count"},
            },
        )

        answer = answer_question("show them", 1, conversation_id="mismatch")

        self.assertEqual(answer["sources"], [])

    def test_reindex_preserves_vectors_and_sheet_locations(self) -> None:
        result = self.upload_workbook(
            [("First", [["Code", "Description"], ["A-1", "ordinary"]], "visible"), ("Final", [["Code", "Description"], ["Z-9", "FINAL_ONLY"]], "visible")],
            "projects.xlsx",
        )
        document_id = int(result["document_id"])
        with database.get_connection() as connection:
            before = connection.execute(
                "SELECT vector_point_id FROM chunks WHERE document_id = ? ORDER BY chunk_index",
                (document_id,),
            ).fetchall()
        vectors = {str(row["vector_point_id"]): [0.0, 1.0] + [0.0] * 382 for row in before}

        class RecordingStore:
            def __init__(self) -> None:
                self.batches = []

            def get_vectors(self, point_ids):
                return {point_id: vectors[point_id] for point_id in point_ids if point_id in vectors}

            def upsert_chunks(self, points):
                self.batches.append(points)

        store = RecordingStore()
        with patch.object(structured_ingestion, "get_vector_store", return_value=store), patch.object(
            structured_ingestion,
            "create_embeddings",
            return_value=[[1.0, 0.0] + [0.0] * 382 for _ in range(len(before))],
        ):
            status = structured_ingestion.reindex_existing_spreadsheet_document(
                document_id=document_id,
                owner_id=1,
                organization_id=ORG,
            )

        with patch.object(vector_search, "create_embeddings", return_value=[[1.0, 0.0] + [0.0] * 382]):
            matches = vector_search.search_chunks("FINAL_ONLY", owner_id=1, document_id=document_id, limit=1)

        self.assertEqual(status.status, "completed")
        self.assertEqual(len(store.batches), 1)
        self.assertEqual(matches[0]["sheet_name"], "Final")

    def test_time_matrix_pairs_in_out_rows_and_filters_working_duration(self) -> None:
        scope = WorkbookScope(
            document_id=7,
            version_id=3,
            filename="attendance.xlsx",
            sheet_names=["April"],
            schema={},
            rows=[
                RowRecord("April", 4, {"Employee Name": "Night Worker", "Employee Number": 99, "Type": "IN", "Mon / 1st": "22:00:00", "Tue / 2nd": "09:00:00"}),
                RowRecord("April", 5, {"Employee Name": "Night Worker", "Employee Number": 99, "Type": "OUT", "Mon / 1st": "08:30:00", "Tue / 2nd": "18:30:00"}),
                RowRecord("April", 6, {"Employee Name": "Night Worker", "Employee Number": 99, "Type": "Total W.hrs", "Mon / 1st": "10:30:00", "Tue / 2nd": "09:30:00"}),
            ],
        )
        result = _answer_time_matrix(
            scope,
            "List all employees who worked more than 10 hours on any day, including employee name, date, IN time, OUT time and total working hours.",
        )

        self.assertIsNotNone(result)
        answer = str(result["answer"])
        self.assertIn("| Night Worker | Mon / 1st | 22:00:00 | 08:30:00 | 10:30:00 |", answer)
        self.assertNotIn("Tue / 2nd", answer)
        self.assertEqual(result["_context"]["numeric_filter"]["column"], "Total Working Hours")


if __name__ == "__main__":
    unittest.main()
